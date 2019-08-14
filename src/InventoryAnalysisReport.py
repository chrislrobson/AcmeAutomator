"""************************************************************************************
CLASS: InventoryAnalysisReport
Author: Christopher Robson
Copyright by:  Christopher Robson
Copyright date: 01Nov2017
This software is not FREE!  Use or destribution of the software system and its
subsystem modules, libraries, configuration file and "seed" file without the
express permission of the author is strictly PROHIBITED!
FUNCTION:  Module generates an Excel formatted inventory report.
***********************************************************************************"""
"""
LIBRARIES:  Python libraries
"""
import xmltodict
from collections import OrderedDict
from openpyxl import Workbook
import ntpath
import datetime
import ast
"""
LIBRARIES:  Home grown specific libraries
"""
from Globals import *
from Utility import Utility
import SearchListDictionary
from Exceptions import XMLKeyError
"""
"""
COLUMN_MAP = {
  "name": 1,
  "version": 2,
  "part-number": 3,
  "serial-number": 4,
  "description": 5,
  "model": 6,
  "disk-size": 7,
  "clei-code": 8,
  "model-number": 8,
  "product": 9,
  "product-number": 10,
  "vendor": 11,
  "die-rev": 12,
  "pcb-rev": 13,
  "mfr-id": 14,
}
"""
CLASS: OrderedDictionary:
DESCRIPTION: 
INPUT: 
OUTPUT: 
"""
class OrderedDictionary:
  "Ordered Dictionary"
  """"""
  def __init__( self ):
    self.name = self.__class__.__name__
  """"""
  def run( self, ws, key, value, rowCnt ):
    self.ws = ws
    self.key = key
    self.value = value
    self.rowCnt = rowCnt
    if type( self.value ) == OrderedDict:
      for self.key, self.value in self.value.items( ):
        self.rowCnt = OrderedDictionary().run( self.ws, self.key, self.value, self.rowCnt )
    elif type( self.value ) is list:
      for self.value in self.value:
        self.rowCnt = StringProcessor( ).run( self.ws, self.key, self.value, self.rowCnt )
    else:
      ## DEBUG print( "KEY: {}, VALUE: {}".format( self.key, self.value ) )
      if key == "name":
        self.rowCnt += 1
      ws.cell( row = self.rowCnt, column = COLUMN_MAP[self.key] ).value = self.value
    return( self.rowCnt )
"""
CLASS: StringProcessor:
DESCRIPTION: 
INPUT: 
OUTPUT: 
"""
class StringProcessor:
  "String Processor"
  """"""
  def __init__( self ):
    self.name = self.__class__.__name__
  """"""
  def run( self, ws, key, value, rowCnt  ):
    self.ws = ws
    self.key = key
    self.value = value
    self.rowCnt = rowCnt
    if type( self.value ) == OrderedDict:
      for self.key, self.value in self.value.items( ):
        self.rowCnt = OrderedDictionary( ).run( self.ws, self.key, self.value, self.rowCnt )
    elif type( self.value ) is list:
      for self.value in self.value:
        self.rowCnt = StringProcessor().run( self.ws, self.key, self.value, self.rowCnt )
    else:
      ## DEBUG print( "KEY: {}, VALUE: {}".format( self.key, self.value ) )
      if self.key == "name":
        self.rowCnt += 1
      ws.cell( row = self.rowCnt, column = COLUMN_MAP[self.key] ).value = self.value
    return( self.rowCnt )
"""
CLASS: JuniperChassisXMLDataParser:
DESCRIPTION: 
INPUT: 
OUTPUT: 
"""
class JuniperChassisXMLDataParser:
  "XML Juniper Chassis Data Parser"
  chassis = ""
  modules = ""
  """"""
  def __init__(self, parent = None):
    self.parent = parent
    self.name = self.__class__.__name__
  """"""
  def juniper_chassis_xml_data_parser( self, dictionary, ws, received_data = None ):
    self.dictionary = dictionary
    self.ws = ws
    self.ws.cell(row=1, column=1).value = "Item Name"
    self.ws.cell(row=1, column=2).value = "Version Number"
    self.ws.cell(row=1, column=3).value = "Part Number"
    self.ws.cell(row=1, column=4).value = "Serial Number"
    self.ws.cell(row=1, column=5).value = "Description"
    self.ws.cell(row=1, column=6).value = "Model Number"
    self.ws.cell(row=1, column=7).value = "Disk Size"
    self.ws.cell(row=1, column=8).value = "Clei Code"
    self.ws.cell(row=1, column=9).value = "Module Number"
    self.ws.cell(row=1, column=10).value = "Product"
    self.ws.cell(row=1, column=11).value = "Product Number"
    self.ws.cell(row=1, column=12).value = "Vendor"
    self.ws.cell(row=1, column=13).value = "Die Rev"
    self.ws.cell(row=1, column=14).value = "PCB Rev"
    self.ws.cell(row=1, column=15).value = "MFR ID"
    if not received_data:
      return( None )
    self.dictionary_data = []
    try:
      self.dictionary_data = xmltodict.parse( received_data )
    except Exception as error:
      self.message = "{{{}{}: {}{}}}".format(Globals.RED_MESSAGE, self.name, error, Globals.SPAN_END_MESSAGE)
      self.dictionary['loggerwidget'].emit(self.message)
      raise Exception
    try:
      obj = self.dictionary_data["rpc-reply"]['chassis-inventory']
    except Exception as error:
      self.message = "{{{}{}: {}{}}}".format(Globals.RED_MESSAGE, self.name, error, Globals.SPAN_END_MESSAGE)
      self.dictionary['loggerwidget'].emit(self.message)
      raise Exception
    root_elements = obj["chassis"] if type(obj) == OrderedDict else [obj["chassis"]]
    self.ws.cell(row=2, column=1).value = root_elements['name']
    self.ws.cell(row=2, column=4).value = root_elements['serial-number']
    self.ws.cell(row=2, column=5).value  = root_elements['description']
    ###DEBUG print( "CHASSIS -> Name: {}, Serial Number: {}, Description: {}".format( root_elements['name'], root_elements['serial-number'],root_elements['description']  ) )
    root_elements = root_elements["chassis-module"] if type(obj) == OrderedDict else [obj["chassis-module"]]
    self.rowCnt = 3
    try:
      for self.element in root_elements:
        for self.key, self.value in self.element.items():
          if type( self.value ) == OrderedDict:
            self.rowCnt = OrderedDictionary().run( self.ws, self.key, self.value, self.rowCnt )
          elif type( self.value ) is list:
            self.rowCnt = StringProcessor().run( self.ws, self.key, self.value, self.rowCnt )
          else:
            ## DEBUG print( "KEY: {}, VALUE: {}".format( self.key, self.value ) )
            if self.key == "name":
              self.rowCnt += 1
            self.ws.cell( row = self.rowCnt, column = COLUMN_MAP[self.key] ).value = self.value
    except Exception as error:
      self.message = "{{{}{}: {}{}}}".format(Globals.RED_MESSAGE, self.name, error, Globals.SPAN_END_MESSAGE)
      self.dictionary['loggerwidget'].emit(self.message)
      raise Exception
    return()
"""
CLASS: InventoryAnalysisReport:
DESCRIPTION: Analysis show chassis hardware data 
INPUT: Data created from the AnalyzeFormattedInventoryData or AnalyzeInventoryData classes
OUTPUT: File containing data ready for Excel report generator
"""
class InventoryAnalysisReport:
  "Inventory Analysis Report"
  """"""
  def __init__(self, parent = None):
    self.parent = parent
    self.name = self.__class__.__name__
    self.search_item_sheet_row = 1
    self.search_item_sheet_column = 1
    self.item_row_no = 1
    self.item_column_no = 1
    self.ToC_initialized = 0
    """
    search_list is built by the Utility class to seed words to find
    for building the Excel table of content
    """
    self.search_list = []
    self.is_search_item = Utility(SearchListDictionary)
    """
    """
    self.testcase_number = ""
    self.testcase_title = ""
    self.inventory_search_data = ""
    self.inventory_data = ""
  """
   CLASS: execute:
   DESCRIPTION: Create "show chassis hardware detail" inventory analysis data entry point.
                Determines input data to process and passes this data onto processing methods.
   INPUT: "show chassis" captured data
   OUTPUT: data ready for generating report data
   """
  def execute( self, dictionary = None, descriptor = None, data = None, decode = None):
    self.dictionary = dictionary
    self.errorRowCnt = 1
    self.table_of_content_list = []
    """
    Setup Excel workbook
    """
    self.wb = Workbook()
    """
    Remove the auto-generated first sheet, a more specific one will be created later
    """
    self.wb.remove(self.wb[self.wb.active.title])
    """
    Built table of contents first
    """
    self.wsSearchItems = self.wb.create_sheet(title="Table of Contents")
    self.wsSearchItems.sheet_properties.tabColor = "1072EE"
    self.wsSearchItems.cell(row=1, column=1).value = "Keyword"
    self.wsSearchItems.cell(row=1, column=2).value = "Sheet Name"
    self.wsSearchItems.cell(row=1, column=3).value = "Sheet Row"
    self.wsSearchItems.cell(row=1, column=4).value = "Sheet Column"
    """
    Setup error report sheet
    """
    self.ws_errors = self.wb.create_sheet(title = "Files with errors")
    """
    Process inventory data now
    """
    if self.dictionary["verbose"]:
      self.message = "{} started at: {}.".format(self.name, datetime.datetime.now().strftime("%d%b%Y-%H%M-%S"))
      self.dictionary['loggerwidget'].emit(self.message)
    self.message = "Building Inventory Analysis Report Data: {}{}.".format(self.dictionary["savepath"], self.dictionary["filename"])
    self.dictionary['loggerwidget'].emit(self.message)
    """
    Get the file where the report data is to be saved
    """
    try:
      self.analsis_report_filename = "{}{}{}{}.xlsx".format(self.dictionary["relativepath"], self.dictionary["savepath"], self.dictionary["filename"], self.dictionary["datetime"])
    except Exception as error:
      self.message = "{{{}{}: {} {}{}}}".format(Globals.RED_MESSAGE, self.name, error, self.dictionary["filename"], Globals.SPAN_END_MESSAGE)
      self.dictionary['loggerwidget'].emit(self.message)
      raise Exception
    """
    Get the file containing the xml formated collected data
    """
    try:
      self.xml_data_filename = "{}{}{}{}".format(self.dictionary["relativepath"], self.dictionary["commandpath"], self.dictionary["commands"], self.dictionary["datetime"])
    except Exception as error:
      self.message = "{{{}{}: {} {}{}}}".format(Globals.RED_MESSAGE, self.name, error, self.dictionary["filename"], Globals.SPAN_END_MESSAGE)
      self.dictionary['loggerwidget'].emit(self.message)
      raise Exception
    try:
      with open(self.xml_data_filename, 'r') as self.xml_data_FD:
        self.testcase_found = False
        for self.xml_data in self.xml_data_FD:
          if self.xml_data.startswith("</testcasenumber>"):
            self.testcase_found = False
            continue
          elif self.xml_data.startswith("<testcasenumber "):
            self.testcase_found = True
            self.testcase_number = self.xml_data.split()[1].split(">")[0]
            if not self.testcase_number:
              self.message = "{{{}{}: {} does not have a testcase number{}}}".format(Globals.RED_MESSAGE, self.name, self.xml_data_filename, Globals.SPAN_END_MESSAGE)
              self.dictionary['loggerwidget'].emit(self.message)
              raise Exception
            continue
          elif self.xml_data.startswith("<testcasetitle>") and self.testcase_found:
            self.testcase_title = self.xml_data.split("<testcasetitle>")[1].split("</testcasetitle>")[0]
            continue
          elif self.xml_data.startswith("<searchseedfile>") and self.testcase_found:
            self.search_seed_filename = self.xml_data.split("<searchseedfile>")[1].split("</searchseedfile>")[0]
            try:
              with open(self.search_seed_filename, 'r') as self.search_file_FD:
                for self.search_list_data in self.search_file_FD:
                  self.is_search_item.set_prompt_string( self.search_list_data.split( "\n" )[0], 1 )
            except Exception as error:
              self.message = "{{{}{}: {} {} ToC will not be created{}}}".format(Globals.BLUE_MESSAGE, self.name, error, self.search_seed_filename, Globals.SPAN_END_MESSAGE)
              self.dictionary['loggerwidget'].emit(self.message)
            self.search_file_FD.close()
            continue
          elif self.xml_data.startswith("<rpc-reply xmlns:") and self.testcase_found:
            try:
              # notes causes issue with name length use analysis seed file title instead --> self.worksheet_name = ntpath.basename(self.analyze_element[self.inventory_data])
              self.worksheet_name = self.testcase_title
              self.inventory_data = self.xml_data
              self.new_xml_data = ""
              while(not self.new_xml_data.startswith("</rpc-reply>")):
                self.new_xml_data = self.xml_data_FD.readline()
                if self.new_xml_data.startswith("<testcasenumber ") or self.new_xml_data.startswith("</testcasenumber>"):
                  self.message = "{{{}{}: malformed XML file - {}{}}}".format(Globals.BLUE_MESSAGE, self.name, self.xml_data_filename, Globals.SPAN_END_MESSAGE)
                  self.dictionary['loggerwidget'].emit(self.message)
                  raise Exception
                else:
                  self.inventory_data += self.new_xml_data
              self.testcase_found = False
              self.ToC_initialized = 0
              self.rowCnt = 1  # First row is the title row
              """
              Create an Excel workbook with the file name of: "Inventory-datetime.xlsx"
              """
              self.ws = self.wb.create_sheet( title = self.worksheet_name )
              self.ws.sheet_properties.tabColor = "1072BA"
              try:
                JuniperChassisXMLDataParser().juniper_chassis_xml_data_parser( self.dictionary, self.ws, self.inventory_data )
              except XMLKeyError as error:
                self.ws_errors.cell( row = self.errorRowCnt, column = 1 ).value = self.inventory_data
                self.errorRowCnt += 1
                self.message = "{{{}{}: {} - {}{}}}".format(Globals.BLUE_MESSAGE, self.name, error, self.inventory_data, Globals.SPAN_END_MESSAGE)
                self.dictionary['loggerwidget'].emit(self.message)
                continue
              try:
                for self.row in self.ws.iter_rows():
                  for self.cell in self.row:
                    if isinstance( self.cell.value, str ):
                      self.search_item_dict = self.is_search_item.find_prompt(self.cell.value, self.is_search_item.prompt_automaton)
                      if self.search_item_dict:
                        self.table_of_content_list.append((self.cell.value, self.worksheet_name, self.cell.row, self.cell.column))
              except Exception as error:
                self.message = "{{{}{}: {}{}}}".format(Globals.RED_MESSAGE, self.name, error, Globals.SPAN_END_MESSAGE)
                self.dictionary['loggerwidget'].emit(self.message)
                raise Exception
            except Exception as error:
              self.message = "{{{}{}: {}{}}}".format(Globals.RED_MESSAGE, self.name, error, Globals.SPAN_END_MESSAGE)
              self.dictionary['loggerwidget'].emit(self.message)
              raise Exception
    except Exception as error:
      self.message = "{{{}{}: {} {} {}}}".format(Globals.RED_MESSAGE, self.name, error, self.xml_data_filename, Globals.SPAN_END_MESSAGE)
      self.dictionary['loggerwidget'].emit(self.message)
      raise Exception

    # todo-debug for self.toc_key, self.ws_page, self.row, self.column in self.table_of_content_list:
    # todo-debug   ###DEBUG print( "Row: {}, Column: {}; Table Key: {}, Worksheet: {}.".
    # todo-debug          format( self.row, self.column, self.toc_key, self.ws_page ) )

    """
    Build a Table of Contents of keywords found in the worksheets just built.
    """
    self.rowCnt = 2
    self.colCnt = 1
    self.ToC_item = []
    for self.toc_key_common, self.ws_page, self.row, self.column in self.table_of_content_list:
      self.flagit = False
      for self.toc_key_item, self.ws_page_item, self.row_item, self.column_item in self.ToC_item:
        if self.toc_key_item == self.toc_key_common and self.ws_page_item == self.ws_page:
          self.flagit = True
      if not self.flagit:
        self.wsSearchItems.cell(row=self.rowCnt, column=self.colCnt).value = self.toc_key_common
        for self.toc_key, self.ws_page, self.row, self.column in self.table_of_content_list:
          if self.toc_key_common == self.toc_key:
            self.wsSearchItems.cell(row=self.rowCnt, column=self.colCnt + 1).value = self.ws_page
            self.wsSearchItems.cell(row=self.rowCnt, column=self.colCnt + 2).value = self.row
            self.wsSearchItems.cell(row=self.rowCnt, column=self.colCnt + 3).value = self.column
            self.rowCnt += 1
            self.ToC_item.append( ( self.toc_key, self.ws_page, self.row, self.column ) )
    """
    Save the worksheets just built.
    """
    try:
      self.wb.save(self.analsis_report_filename)
    except Exception as error:
      self.message = "{{{}{}: {}{}}}".format(Globals.RED_MESSAGE, self.name, error, Globals.SPAN_END_MESSAGE)
      self.dictionary['loggerwidget'].emit(self.message)
      raise Exception
    return()
"""
End of File
"""

