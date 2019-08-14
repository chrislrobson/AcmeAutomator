#####################################################################################################################
# Python Qt5 Testbed Tester Create Master Seed From Excel Processor
# Author: Christopher Robson
# Copyright by:  Christopher Robson
# Copyright date: 01Jan2016
# This software is not FREE!  Use or destribution of the software system and its
# subsystem modules, libraries, configuration file and "seed" file without the
# express permission of the author is strictly PROHIBITED!
# FUNCTION: This class create a master seed file from an excel workbook
#####################################################################################################################
import importlib
from openpyxl import load_workbook
from PyQt5 import QtCore
#-------------------------------------------------------------------------------------------------------------------
# Home grown stuff
#-------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------
class ExcelWorkbookProcessor( QtCore.QThread ):
  " Excel Workbook Processor"
  #-----------------------------------------------------------------------------------------------------------------
  def __init__(self, parent = None ):
    super( ExcelWorkbookProcessor, self).__init__( parent )
    self.parent = parent
    self.name = " Create Master Seed from Excel Processor"
  #-----------------------------------------------------------------------------------------------------------------
  # Process Excel Workbook
  #-----------------------------------------------------------------------------------------------------------------
  def process_workbook( self ):
    # FIXME change how this file is handled later !!!!
    # os.remove( "./test-master-profiles.prf" )
    self.filename = "./test-template-master-profile.xlsx"
    self.workbook = load_workbook( self.filename )
    self.parent.processor_message_signal.emit( "Workbook {} spreadsheets:".format( self.filename ))
    for self.sheetname in self.workbook.get_sheet_names():
      self.parent.processor_message_signal.emit( "  " + self.sheetname )
      self.sheet_data = self.workbook[self.sheetname]
      self.row_count = self.sheet_data.max_row - 5 # dont include the BASELINE
      self.cell = iter( self.sheet_data["A"] )
      while( True ):
        self.section = []
        try:
          self.cell_data = next( self.cell )
          try:
            self.section, self.cell = self.build_configuration_section( self.cell )
          except:
            pass
          for self.line in self.section:
            self.write_cell_data_to_confguration_file( self.line )
          self.write_cell_data_to_confguration_file( "!" )
        except StopIteration:
          break
    self.read_data_from_confguration_file()
    return()
  #-----------------------------------------------------------------------------------------------------------------
  def write_cell_data_to_confguration_file( self, data ):
    self.data = data
    try:
      with open( "./test-master-profiles.prf", 'a+' ) as self.masterFD:
        # FIXME WHY AM I DOING THIS if self.data[0].isupper() or self.data[0].isdigit():
        # FIXME WHY AM I DOING THIS   self.data = "! " + self.data
        self.masterFD.write( self.data + "\n" )
        self.masterFD.close()
    except OSError as error:
      raise Exception( "{}: File error: {}.".format( self.name, error.args[1] ) )
    return()
  #-----------------------------------------------------------------------------------------------------------------
  def read_data_from_confguration_file( self ):
    self.parent.processor_message_signal.emit( "{}".format( "\n-------------------------------------------------"
                                                            "-------------------------------------------------" ) )
    self.parent.processor_message_signal.emit( "Configuration file {} data".format( "./test-master-profiles.prf" ) )
    self.parent.processor_message_signal.emit( "{}".format( "-------------------------------------------------"
                                                            "-------------------------------------------------" ) )
    try:
      with open( "./test-master-profiles.prf", 'r' ) as self.masterFD:
        self.data_read = self.masterFD.read()
        self.parent.processor_message_signal.emit( self.data_read )
    except OSError as error:
      raise Exception( "{}: File error: {}.".format( self.name, error.args[1] ) )
    self.masterFD.close()
    return()
  #-----------------------------------------------------------------------------------------------------------------
  #-----------------------------------------------------------------------------------------------------------------
  def build_configuration_section( self, cell ):
    self.section = self.cell_data.value.split()[0]
    try:
      self.module = importlib.import_module("ExcelWorkbookProcessor")
      self.module_class = getattr( self.module, self.section )
      self.module_method = self.module_class( self )
      self.config_section, self.cell = self.module_method.execute( self.cell )
    except Exception as error:
      pass
      # FIXME REMOVE MAYBE ? self.message_str = Globals.RED_MESSAGE + \
      # FIXME REMOVE MAYBE ?                    "ANANLYSIS: File I/O error with {} error reported: {}.". \
      # FIXME REMOVE MAYBE ?                      format(list(top_dict.keys())[0], error.args[0]) + \
      # FIXME REMOVE MAYBE ?                    Globals.SPAN_END_MESSAGE
      # FIXME REMOVE MAYBE ? self.ggparent.logger_message_signal.emit(self.message_str)
      # FIXME REMOVE MAYBE ? self.cmd_list[SeedCommandDictionaryProcessor.archivefilenameFD].close()
    return( self.config_section, self.cell )
####################################################################################################################
