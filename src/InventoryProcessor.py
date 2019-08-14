##################################################################################################################
#  Inventory Processor
#
# This script will SSH into each of the routers listed
# in the "networkDeviceProfiles" file and issuing a 
# "show chassis".  It takes this data and pushes it to an 
# Excel file "Inventory-[datetime].xlsl".  This file contains
# seperate worksheets for each device, the IP address of
# the device is used as the worksheet name.
# TO control this program the following file is passed which
# contains the profile of each network device to pull from.
#  The file's profiles are in the following format:
# ---- juniper_10_168_1_199 = {'device_type':'juniper',
#      'ip':'10.168.1.199','username':'root',
#      'password':'junos01','secret':'','port':22,
#      'verbose':False,}
# 
##################################################################################################################
from openpyxl import Workbook
from PyQt5 import QtCore
import ntpath
import datetime
#
# Home grown stuff
#
from Globals import *
#
class InventoryProcessor( QtCore.QThread ):
  " Inventory Spreadsheet Processor"
  #
  def __init__(self, parent = None ):
    super( InventoryProcessor, self).__init__( parent )
    self.parent = parent
    self.name = " Inventory Spreadsheet Processor"
  #
  # Process Device Inventory
  #
  def process_inventory( self ):
    self.xlsx_filename = ""
    #
    self.wb = Workbook()
    self.wb.remove_sheet(self.wb.active)
    self.data = None
    #
    try:
      for self.filename in Globals().get_show_file_list():
        #
        self.ws = self.wb.create_sheet(title=ntpath.basename(self.filename))
        self.ws.sheet_properties.tabColor = "1072BA"
        #
        self.ws["A1"] = "Item"
        self.ws["B1"] = "Version"
        self.ws["C1"] = "Part No."
        self.ws["D1"] = "CLEI Code"
        self.ws["E1"] = "FRU Module No."
        self.rowCnt = 1  # First row is the title row
        with open( self.filename, 'r' ) as self.file_FD:
          for self.line in self.file_FD:
            if "Hardware" in self.line:
              pass
            elif "Item" in self.line:
              pass
            elif "root>" in self.line:
              pass
            elif "@" in self.line:
              pass
            elif self.line.startswith( "DUT" ) or \
                 self.line.startswith( "#" ) or \
                 self.line.startswith( "!" ) or \
                 self.line.startswith( "\n" ) or \
                 self.line.startswith( "show " ) or \
                 self.line.startswith( "File date/time:" ):
              pass
            elif self.line.startswith( "<rpc-reply" ):
              #DEBUG self.parent.logger_message_signal.emit( "XML file found, skippng this spreadsheet." )
              #
              # Remove sheet created
              #
              self.wb.remove_sheet( self.ws )
              break
            else:
              #
              # Worksheet cell update section.  Data pull
              # from a router "show chassis" command is pushed
              # into an appropriate Excel cell.
              #
              self.data = self.line.split()
              self.wordCnt = len( self.data )
              self.alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
              for self.rc in range( self.wordCnt ):
                a = self.alphabet[self.rc]
                i = self.rc
                self.ws[self.alphabet[self.rc] + str(self.rowCnt)] = self.data[self.rc]
              self.rowCnt += 1
    except Exception as error:
      self.message_str = Globals.RED_MESSAGE + \
                    "INVENTORYSPREADSHEETPROCESSOR: File I/O error, inventory process has aborted!" + \
                    Globals.SPAN_END_MESSAGE
      self.parent.logger_message_signal.emit(self.message_str)
    #
    self.xlsx_filename = Globals.inventory_directory + \
                         "Inventory" + datetime.datetime.now().strftime( "-%d%b%Y-%H%M-%S" ) + \
                         ".xlsx"
    self.wb.save( self.xlsx_filename )
    return()
####################################################################################################################
